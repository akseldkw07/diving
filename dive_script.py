"""Append new Garmin dive activities to the Scuba Log spreadsheet.

Handles Garmin's English and Spanish CSV exports transparently. Both end up
parsed into the same canonical English column set before the script writes to
the workbook.

Garmin's English export has been observed in three variants:
  (a) clean 27-column: "Total Ascent" / "Total Descent" omitted from header
      AND from each data row (no fix needed).
  (b) broken 27/29: header lists only 27 names (with 2 trailing empty fields)
      but each data row has 29 fields including ascent/descent values. We patch
      the header by inserting the two missing names after "Max Speed".
  (c) intact 29-column: matches the Spanish layout.

The Spanish export is always intact; we translate column names to English.
"""

import csv
import io

import pandas as pd
from openpyxl import load_workbook as lw

SPANISH_TO_ENGLISH = {
    "Tipo de actividad": "Activity Type",
    "Fecha": "Date",
    "Favorito": "Favorite",
    "Título": "Title",
    "Distancia": "Distance",
    "Calorías": "Calories",
    "Tiempo": "Time",
    "Frecuencia cardiaca media": "Avg HR",
    "FC máxima": "Max HR",
    "Velocidad media": "Avg Speed",
    "Velocidad máxima": "Max Speed",
    "Ascenso total": "Total Ascent",
    "Descenso total": "Total Descent",
    "Pasos": "Steps",
    "Profundidad máxima": "Max Depth",
    "Tiempo de inmersión": "Dive Time",
    "Temperatura mínima del agua": "Min Water Temp",
    "Tipo de gas": "Gas Type",
    "Intervalo en superficie": "Surface Interval",
    "Descompresión": "Decompression",
    "Tipo de agua": "Water Type",
    "Mejor tiempo de vuelta": "Best Lap Time",
    "Número de vueltas": "Number of Laps",
    "Temperatura máxima": "Max Temp",
    "Tiempo en movimiento": "Moving Time",
    "Tiempo transcurrido": "Elapsed Time",
    "Altura mínima": "Min Elevation",
    "Altura máxima": "Max Elevation",
}


def _normalize_surface_interval(x: str) -> str:
    """Garmin sometimes leaves Surface Interval blank or as `--`. Treat both as zero."""
    if not isinstance(x, str):
        return "0:00:00"
    x = x.strip()
    if x == "" or x == "--":
        return "0:00:00"
    if x.count(":") < 2:
        return "0:" + x
    return x


def _load_dives(path: str) -> pd.DataFrame:
    """Read a Garmin dive activities CSV and return a DataFrame with canonical
    English column names, handling Spanish exports and broken English headers."""
    with open(path, "rb") as f:
        raw = f.read().decode("utf-8-sig")

    rows = list(csv.reader(io.StringIO(raw)))
    if not rows:
        raise ValueError(f"Empty CSV: {path}")

    header, data_rows = rows[0], rows[1:]
    is_spanish = header[0].startswith("Tipo de actividad")

    # English-only fixup: if data rows have 2 more fields than the header has
    # named columns, Garmin dropped "Total Ascent"/"Total Descent" from the
    # header. Patch them in after "Max Speed".
    if not is_spanish and data_rows:
        named_count = sum(1 for h in header if h.strip())
        max_data_cols = max(len(r) for r in data_rows)
        if max_data_cols == named_count + 2 and "Total Ascent" not in header and "Max Speed" in header:
            i = header.index("Max Speed")
            tail = [h for h in header[i + 1 :] if h.strip()]
            header = header[: i + 1] + ["Total Ascent", "Total Descent"] + tail

    # Rebuild the CSV in memory with the (possibly patched) header and let
    # pandas do the rest of the parsing.
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    writer.writerows(data_rows)
    buf.seek(0)
    df = pd.read_csv(buf)

    if is_spanish:
        df = df.rename(columns=SPANISH_TO_ENGLISH)

    return df


def main():
    dive_csv_path = "/Users/Akseldkw/Desktop/Aksel/Diving/Activities.csv"
    dive_log_path = "/Users/Akseldkw/Desktop/Aksel/Diving/Scuba_Log.xlsx"
    wb = lw(dive_log_path)
    ws = wb["Log"]
    print(ws.max_row)
    print(len(ws["A"]))
    print(ws[f"A{ws.max_row}"].value)
    index_start = ws[f"A{ws.max_row}"].value + 1
    new_dives = _load_dives(dive_csv_path)

    columns = [
        "Date",
        "0",
        "1",
        "2",
        "3",
        "4",
        "Max Depth",
        "Min Water Temp",
        "5",
        "Dive Time In",
        "Gas Type",
        "Time",
        "Surface Interval",
        "7",
        "Avg HR",
        "Max HR",
    ]
    new_dives["Date"] = pd.to_datetime(new_dives["Date"])
    new_dives.sort_values(by=["Date"], ascending=True, inplace=True)

    new_dives["Surface Interval"] = pd.to_timedelta(
        new_dives["Surface Interval"].apply(_normalize_surface_interval)
    ).astype(str)
    new_dives["Dive Time In"] = new_dives["Date"].dt.strftime("%I:%M %p")
    new_dives["Date"] = new_dives["Date"].dt.strftime("%-m/%-d/%Y")
    new_dives["Time"] = (pd.to_timedelta(new_dives["Time"]).dt.total_seconds() / 60).round(0)

    new_dives.reset_index(inplace=True)
    for index in range(8):
        new_dives[str(index)] = ""

    new_dives = new_dives[columns]
    new_dives.index += index_start
    new_dives.reset_index(inplace=True)
    print(new_dives.info())
    print(new_dives)

    with pd.ExcelWriter(path=dive_log_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        start_row = len(ws["A"])
        new_dives.to_excel(writer, sheet_name="Log", index=False, startrow=start_row, header=False)


if __name__ == "__main__":
    main()
